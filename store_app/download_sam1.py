import re

import pandas as pd
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .inventory_utils import add_inventory_with_serials
from .models import Product, StockMovement, Vendor


HEADER_ALIASES = {
    "vendor_id": ["vendor_id", "vendor id", "vendor_i", "vendor"],
    "prefix_code": ["prefix_code", "prefix code", "prefix"],
    "sku": ["sku", "product sku"],
    "name": ["name", "item name", "product name"],
    "size": ["size"],
    "color": ["color", "colour"],
    "material": ["material"],
    "serial": ["serial", "serial no", "serial number"],
    "unit_purchase_price": ["unit_purchase_price", "unit purchase price", "unit_pur", "unit_purc", "price"],
    "desc": ["desc", "description"],
    "weight_before": ["weight_before", "weight before", "weight_b"],
    "weight_after": ["weight_after", "weight after", "weight_a"],
    "length": ["length"],
    "unit": ["unit"],
    "width": ["width"],
    "height": ["height"],
    "quantity": ["quantity", "qty", "stock", "qnty"],
}


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_float(value):
    text = clean_text(value).replace(",", "")
    try:
        return float(text) if text else 0
    except ValueError:
        return 0


def clean_int(value, default=0):
    text = clean_text(value).replace(",", "")
    try:
        return int(float(text)) if text else default
    except ValueError:
        return default


def clean_optional_int(value):
    text = clean_text(value).replace(",", "")
    try:
        return int(float(text)) if text else None
    except ValueError:
        return None


def normalize_header(value):
    value = clean_text(value).replace("\n", " ").replace("_", " ")
    return " ".join(value.split()).lower()


def apply_header_aliases(df):
    rename_map = {}

    for original_col in df.columns:
        normalized_col = normalize_header(original_col)

        for target, aliases in HEADER_ALIASES.items():
            all_names = [normalize_header(target)] + [normalize_header(alias) for alias in aliases]
            if normalized_col in all_names:
                rename_map[original_col] = target
                break

    return df.rename(columns=rename_map)


def slug_code(value):
    value = clean_text(value).upper()
    value = re.sub(r"[^A-Z0-9]+", "-", value).strip("-")
    return value


def generate_sku(prefix_code):
    prefix_code = slug_code(prefix_code)

    last_product = (
        Product.objects
        .filter(prefix_code__iexact=prefix_code)
        .exclude(sku__isnull=True)
        .exclude(sku="")
        .order_by("-id")
        .first()
    )

    next_number = 1
    if last_product and last_product.sku:
        last_part = clean_text(last_product.sku).split("-")[-1]
        if last_part.isdigit():
            next_number = int(last_part) + 1

    return f"{prefix_code}-{next_number:05d}"
def download_sample1(request):
    from .models import Vendor
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "vendor_id", "prefix_code", "sku", "name", "size",
        "color", "material", "unit_purchase_price", "desc",
        "weight_before", "weight_after", "length", "unit",
        "width", "height", "quantity"
    ]

    header_fill = PatternFill(start_color="1B175D", end_color="1B175D", fill_type="solid")
    qty_header_fill = PatternFill(start_color="31a950", end_color="31a950", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    qty_col_index = headers.index("quantity") + 1

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.fill = qty_header_fill if col_num == qty_col_index else header_fill

    # ── Column width ───────────────────────────────────────────────
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 15)

    ws.freeze_panes = "A2"

    # ── Sheet 2: Vendors ───────────────────────────────────────────
    ws2 = wb.create_sheet(title="Vendors")
    vendor_headers = ["id", "name", "firm_name", "mobile", "city", "state"]
    vendor_header_fill = PatternFill(start_color="2E2A8A", end_color="2E2A8A", fill_type="solid")

    for col_num, header in enumerate(vendor_headers, 1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = vendor_header_fill

    vendors = Vendor.objects.all().order_by('name')
    for row_num, v in enumerate(vendors, 2):
        row_data = [v.id, v.name or "", v.firm_name or "", v.mobile or "", v.city or "", v.state or ""]
        for col_num, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal='left')

    for col_num, header in enumerate(vendor_headers, 1):
        col_letter = get_column_letter(col_num)
        max_len = len(header)
        for row in ws2.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max_len + 4, 30)

    ws2.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=inventory_sample.xlsx"
    return response
@csrf_exempt
def upload_products1(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    file = request.FILES.get("file")
    if not file:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    try:
        file_name = file.name.lower()

        if file_name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(file)
        elif file_name.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            return JsonResponse(
                {"error": "Only .xlsx, .xls and .csv files are allowed"},
                status=400,
            )

        df = df.dropna(how="all")
        df.columns = [normalize_header(col) for col in df.columns]
        df = apply_header_aliases(df)

        required_columns = ["vendor_id", "prefix_code", "name", "quantity"]
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return JsonResponse(
                {"error": f"Missing columns: {', '.join(missing_columns)}"},
                status=400,
            )

        success_count = 0
        errors = []

        for index, row in df.iterrows():
            row_number = index + 2

            try:
                vendor_id_raw = row.get("vendor_id")
                prefix_code = clean_text(row.get("prefix_code"))
                sku = clean_text(row.get("sku"))
                name = clean_text(row.get("name"))
                quantity = clean_int(row.get("quantity"), 0)

                if not vendor_id_raw or clean_text(vendor_id_raw) == "":
                    errors.append(f"Row {row_number}: vendor_id is required")
                    continue

                if not prefix_code:
                    errors.append(f"Row {row_number}: prefix_code is required")
                    continue

                if not name:
                    errors.append(f"Row {row_number}: name is required")
                    continue

                if quantity <= 0:
                    errors.append(f"Row {row_number}: quantity must be greater than 0")
                    continue

                vendor = Vendor.objects.filter(id=int(float(vendor_id_raw))).first()
                if not vendor:
                    errors.append(f"Row {row_number}: Vendor {vendor_id_raw} not found")
                    continue

                prefix_code = slug_code(prefix_code)
                size = clean_text(row.get("size"))
                color = clean_text(row.get("color"))
                material = clean_text(row.get("material"))
                serial = clean_optional_int(row.get("serial"))

                with transaction.atomic():
                    product = Product.objects.filter(sku=sku).first() if sku else None

                    if not product:
                        product = Product.objects.filter(
                            vendor=vendor,
                            name=name,
                            size=size,
                            color=color,
                            material=material,
                        ).first()

                    if product:
                        product.prefix_code = prefix_code
                        product.name = name
                        product.size = size
                        product.color = color
                        product.material = material
                        product.serial = serial
                        product.unit_purchase_price = clean_float(row.get("unit_purchase_price"))
                        product.desc = clean_text(row.get("desc"))
                        product.weight_before = clean_text(row.get("weight_before"))
                        product.weight_after = clean_text(row.get("weight_after"))
                        product.length = clean_text(row.get("length"))
                        product.unit = clean_text(row.get("unit")) or "piece"
                        product.width = clean_text(row.get("width"))
                        product.height = clean_text(row.get("height"))

                        if not clean_text(product.sku):
                            product.sku = generate_sku(prefix_code)
                        elif not clean_text(product.sku).startswith(prefix_code + "-"):
                            product.sku = generate_sku(prefix_code)

                        product.save()
                    else:
                        product = Product.objects.create(
                            vendor=vendor,
                            prefix_code=prefix_code,
                            name=name,
                            size=size,
                            color=color,
                            material=material,
                            serial=serial,
                            sku=generate_sku(prefix_code),
                            unit_purchase_price=clean_float(row.get("unit_purchase_price")),
                            desc=clean_text(row.get("desc")),
                            weight_before=clean_text(row.get("weight_before")),
                            weight_after=clean_text(row.get("weight_after")),
                            length=clean_text(row.get("length")),
                            unit=clean_text(row.get("unit")) or "piece",
                            width=clean_text(row.get("width")),
                            height=clean_text(row.get("height")),
                        )

                    serials = add_inventory_with_serials(product, quantity)
                    StockMovement.objects.create(
                        product=product,
                        delta=quantity,
                        reason="ADJUST",
                        note=(
                            "Inventory added via Excel. "
                            f"Serials: {serials[0]} to {serials[-1]}"
                        ),
                    )

                success_count += 1

            except Exception as e:
                errors.append(f"Row {row_number}: {str(e)}")

        return JsonResponse({
            "status": "success",
            "uploaded": success_count,
            "errors": errors,
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
